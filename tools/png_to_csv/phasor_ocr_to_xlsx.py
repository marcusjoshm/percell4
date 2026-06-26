"""
phasor_ocr_to_xlsx.py
Extract the data table from Leica Phasor Calibration screenshots (PNG)
and save Name / Channel / Phase / Modulation to an Excel spreadsheet.

------------------------------------------------------------------------
SETUP (one-time)
------------------------------------------------------------------------
1. Install Tesseract OCR (the actual OCR engine):
     Windows: https://github.com/UB-Mannheim/tesseract/wiki  (then add the
              install folder to PATH, or set TESSERACT_CMD below)
     macOS:   brew install tesseract
     Linux:   sudo apt install tesseract-ocr

2. Install the Python packages:
     pip install pytesseract Pillow openpyxl

------------------------------------------------------------------------
USAGE
------------------------------------------------------------------------
   python phasor_ocr_to_xlsx.py image1.png [image2.png ...] -o output.xlsx

If -o is omitted, output goes to phasor_calibration.xlsx in the current
directory. Multiple images are concatenated in order; rows that appear
in more than one image (because of scroll overlap) are de-duplicated.

------------------------------------------------------------------------
EXPECTED ACCURACY
------------------------------------------------------------------------
- Phase, Modulation, and Channel are nearly always perfect (regex-anchored).
- Names are usually clean. Common OCR slip-ups to look for:
    * "FLIMS" or "FLIM5" should be "FLIM 5" (missing space)
    * "Hpep1-1" should be "Hpep1 - 1" (missing spaces around dash)
    * The very top row may be skipped if the table is scrolled and the
      row is half-cut at the top of the image.
  Open the output xlsx and skim the Name column; fix any oddities with
  Excel's Find & Replace before using the data.
"""

import argparse
import re
import sys
from pathlib import Path

from PIL import Image, ImageOps
import pytesseract
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# If tesseract isn't on PATH (common on Windows), set its full path here:
# pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# ---------------------------------------------------------------------------
# OCR + PARSING
# ---------------------------------------------------------------------------

# A data row looks like:  <name>  HyD X 3  ✓  34.84 °  0.9997
# We anchor on the two most reliable tokens — the CHANNEL and the MODULATION
# (always 0.dddd) — and pull the phase NUMBER from between them. The degree
# symbol is ignored entirely: OCR renders it as °, *, ', ` … which used to
# break a strict `\s*°?\s+` separator and silently drop ~3/4 of the rows.
# Numbers (phase + modulation) are what matter; names are best-effort and easy
# to fix by hand, and a row is never dropped just because its name is messy.
CHAN_RE = re.compile(r"HyD\s*([XS])\s*(\d)")
MOD_RE = re.compile(r"0\.\d{3,4}")
PHASE_DOT_RE = re.compile(r"-?\d{1,3}\.\d{1,2}")     # e.g. 38.92
PHASE_NODOT_RE = re.compile(r"-?\d{3,4}")            # e.g. 3892 -> 38.92
# Any line carrying a channel followed by a modulation is treated as a data row.
DATA_ROW_RE = re.compile(r"HyD\s*[XS]\s*\d.*?0\.\d{3,4}")

JUNK_CHARS = set("SVYvLsylWw=.¥√✓/\\|_~-&\"'`*°")


def clean_name(raw: str) -> str:
    """Strip leading OCR garbage that comes from the two checkmark columns."""
    s = raw.strip()
    changed = True
    while changed:
        changed = False
        s2 = re.sub(r"^[^A-Za-z0-9]+", "", s)
        if s2 != s:
            s, changed = s2, True
        m = re.match(r"^([A-Za-z])\s+(?=\S)", s)            # single letter + space
        if m:
            s, changed = s[m.end():], True
            continue
        m = re.match(r"^(\S{2,3})\s+(?=\S)", s)              # 2-3 char junk-only word
        if m and all(c in JUNK_CHARS for c in m.group(1)):
            s, changed = s[m.end():], True
            continue
        m = re.match(r"^([SVYLW])([A-Z][a-z])", s)            # junk-letter glued to word
        if m:
            s, changed = s[1:], True
    # Common per-character OCR slips
    s = re.sub(r"\bOmin\b", "0min", s)
    s = re.sub(r"(\d)ul\b", r"\1uL", s)
    return s.strip()


def extract_rows(image_path: Path):
    """Run OCR on one image, return list of (name, channel, phase, modulation).

    Best-effort: every line that carries a channel + a modulation is emitted as
    a row. The phase/modulation NUMBERS are extracted regardless of the (often
    mis-OCR'd) degree symbol; if the phase or name can't be read cleanly the row
    is still written — phase blank, name blank/partial — and a warning is logged
    so the user can fill those cells in by hand. Rows are never silently dropped.
    """
    img = Image.open(image_path).convert("L")
    img = ImageOps.invert(img)                                # dark UI -> dark-on-light
    img = img.resize((img.width * 2, img.height * 2), Image.LANCZOS)
    text = pytesseract.image_to_string(img, config="--psm 6")

    rows = []
    warnings = []
    for line in text.splitlines():
        if not DATA_ROW_RE.search(line):
            continue
        chan_m = CHAN_RE.search(line)
        mod_m = MOD_RE.search(line, chan_m.end())             # modulation after channel
        if mod_m is None:                                     # shouldn't happen (gated)
            warnings.append(f"no modulation read: {line.strip()}")
            continue

        channel = f"HyD {chan_m.group(1)} {chan_m.group(2)}"
        modulation = float(mod_m.group())

        # Phase = the number between channel and modulation; the degree glyph
        # (°, *, ', …) is just skipped over.
        between = line[chan_m.end():mod_m.start()]
        phase, note = None, ""
        pm = PHASE_DOT_RE.search(between)
        if pm:
            phase = float(pm.group())
        else:
            pm = PHASE_NODOT_RE.search(between)               # OCR dropped the dot
            if pm:
                raw = pm.group()
                sign = -1.0 if raw.startswith("-") else 1.0
                digits = raw.lstrip("-")
                phase = sign * float(f"{digits[:-2]}.{digits[-2:]}")
                note = f"phase reconstructed as {phase} from {raw!r}"
            else:
                note = "phase unreadable (left blank)"

        name = clean_name(line[:chan_m.start()])
        if not name:
            note = (note + "; " if note else "") + "name unreadable (left blank)"

        rows.append((name, channel, phase, modulation))
        if note:
            warnings.append(f"{note}: {line.strip()}")

    if warnings:
        print(f"  ⚠ {image_path.name}: {len(warnings)} row(s) need a manual check "
              f"(numbers were still extracted where possible):", file=sys.stderr)
        for w in warnings:
            print(f"      {w}", file=sys.stderr)
    return rows


def dedupe(rows):
    """Drop exact duplicates while preserving order (handles scroll overlap)."""
    seen = set()
    out = []
    for r in rows:
        if r not in seen:
            seen.add(r)
            out.append(r)
    return out


def normalize_group_names(rows, threshold=0.90):
    """Make sure consecutive same-sample rows have identical Name strings.

    A Leica sample typically spans 2-3 consecutive rows (one per detector
    channel). OCR can produce slightly different name strings on each row
    (e.g. "As + Noco" vs "As +Noco"), which breaks downstream automation
    that joins on Name. Here we group runs of consecutive rows whose names
    are >= `threshold` similar, then force every row in the group to share
    one canonical name (the most frequent; ties broken by longest, which
    usually beats OCR-truncated variants)."""
    from collections import Counter
    from difflib import SequenceMatcher

    if not rows:
        return rows, []

    def digits_only(s):
        return "".join(c for c in s if c.isdigit())

    groups = [[0]]
    for i in range(1, len(rows)):
        prev, cur = rows[i - 1][0], rows[i][0]
        sim = SequenceMatcher(None, prev, cur).ratio()
        # Require both high similarity AND identical digit sequence; this
        # prevents "Hpep1" and "Hpep3" (or "60min" and "80min") from being
        # collapsed since digits usually carry sample-identifying meaning.
        if sim >= threshold and digits_only(prev) == digits_only(cur):
            groups[-1].append(i)
        else:
            groups.append([i])

    out = list(rows)
    changes = []
    for group in groups:
        if len(group) < 2:
            continue
        names = [rows[i][0] for i in group]
        counts = Counter(names)
        top_count = max(counts.values())
        candidates = [n for n, c in counts.items() if c == top_count]
        canonical = max(candidates, key=len)
        for i in group:
            n, ch, p, m = rows[i]
            if n != canonical:
                changes.append((n, canonical))
                out[i] = (canonical, ch, p, m)
    return out, changes


# ---------------------------------------------------------------------------
# XLSX WRITER
# ---------------------------------------------------------------------------

def write_xlsx(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Phasor Calibration"

    headers = ["Name", "Channel", "Phase (°)", "Modulation"]
    ws.append(headers)

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="305496")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, 5):
        c = ws.cell(row=1, column=col)
        c.font, c.fill, c.alignment, c.border = header_font, header_fill, center, border

    body_font = Font(name="Arial")
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.font = body_font
            cell.border = border
            cell.alignment = left if cell.column == 1 else center

    name_w = max((len(r[0]) for r in rows), default=20) + 2
    ws.column_dimensions["A"].width = max(22, name_w)
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=3).number_format = "0.00"
        ws.cell(row=r, column=4).number_format = "0.0000"

    ws.freeze_panes = "A2"
    wb.save(output_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description=__doc__.split("---")[0],
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("images", nargs="+", help="One or more PNG screenshots")
    ap.add_argument("-o", "--output", default="phasor_calibration.xlsx",
                    help="Output xlsx path (default: phasor_calibration.xlsx)")
    ap.add_argument("--print-only", action="store_true",
                    help="Print parsed rows and skip writing the xlsx")
    args = ap.parse_args()

    all_rows = []
    for p in args.images:
        path = Path(p)
        if not path.exists():
            sys.exit(f"File not found: {p}")
        rows = extract_rows(path)
        print(f"  {path.name}: {len(rows)} rows", file=sys.stderr)
        all_rows.extend(rows)

    all_rows = dedupe(all_rows)
    # NB: name normalization is intentionally NOT applied — with best-effort OCR
    # names, forcing a canonical name across a sample pair could overwrite a
    # clean name with a garbled neighbour. Names are left as-read for the user
    # to correct; the numbers are the load-bearing output.
    print(f"Total after dedupe: {len(all_rows)} rows", file=sys.stderr)

    if args.print_only:
        for r in all_rows:
            print(r)
        return

    write_xlsx(all_rows, args.output)
    print(f"Wrote {args.output}", file=sys.stderr)


if __name__ == "__main__":
    main()
